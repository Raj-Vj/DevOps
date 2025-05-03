
import openpyxl
import subprocess
from pathlib import Path

# Path to the Excel file
excel_path = Path("DevOps_Study_Tracker.xlsx")

# Load workbook and worksheet
wb = openpyxl.load_workbook(excel_path)
ws = wb["Daily Learning Plan"]

# Prompt user to update a task
day = int(input("Enter the Day number to update (e.g., 1): "))
status = input("Enter the new status (e.g., Done/In Progress): ")

# Update the status in the sheet
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == day:
        row[4].value = status
        break

# Save the workbook
wb.save(excel_path)
print(f"Day {day} status updated to '{status}'.")

# Git commands to commit and push changes
repo_path = Path.cwd()
commit_message = f"Update: Day {day} marked as {status}"

try:
    subprocess.run(["git", "-C", str(repo_path), "add", str(excel_path)], check=True)
    subprocess.run(["git", "-C", str(repo_path), "commit", "-m", commit_message], check=True)
    subprocess.run(["git", "-C", str(repo_path), "push"], check=True)
    print("Changes pushed to GitHub successfully.")
except subprocess.CalledProcessError as e:
    print("Git error:", e)
